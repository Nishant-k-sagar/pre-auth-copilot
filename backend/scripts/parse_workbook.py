"""
CLI entry point for workbook parsing.

Usage:
    Set-Location backend
    python scripts/parse_workbook.py --input data/preauth_workbook.xlsx --output data/

Imports shared parsing logic from skill.parser_utils.
Writes 4 JSON files: training_cases, patient_data_schema, complex_case, expected_outcome.
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Union

from openpyxl import load_workbook

# Make backend/ importable as a package root
sys.path.insert(0, str(Path(__file__).parent.parent))

from skill.parser_utils import (
    ComplexCaseRow,
    ExpectedOutcome,
    SchemaFieldRow,
    TrainingCaseRow,
    parse_training_cases,
    parse_patient_data_schema,
    parse_complex_case,
    parse_expected_outcome,
    validate_parsed_data,
)


WorkbookFileData = Union[
    List[TrainingCaseRow],
    List[SchemaFieldRow],
    List[ComplexCaseRow],
    ExpectedOutcome,
]


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse pre-auth workbook to JSON")
    parser.add_argument("--input", required=True, help="Path to .xlsx workbook")
    parser.add_argument("--output", required=True, help="Output directory for JSON files")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        print(f"Workbook not found: {input_path}")
        sys.exit(1)

    print(f"Reading: {input_path}")
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    print(f"Sheets:  {workbook.sheetnames}")

    print("\nParsing sheets...")
    training_cases = parse_training_cases(workbook)
    schema_fields = parse_patient_data_schema(workbook)
    complex_case = parse_complex_case(workbook)
    expected_outcome = parse_expected_outcome(workbook)

    files: Dict[str, WorkbookFileData] = {
        "training_cases.json": training_cases,
        "patient_data_schema.json": schema_fields,
        "complex_case.json": complex_case,
        "expected_outcome.json": expected_outcome,
    }

    for filename, data in files.items():
        path = output_dir / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"  Written: {path}")

    ok = validate_parsed_data(training_cases, schema_fields, complex_case, expected_outcome)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
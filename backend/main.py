"""
FastAPI application entry point.

CRITICAL: load_dotenv() is the very first statement — before any other imports.
This ensures MISTRAL_API_KEY is in os.environ when pipeline.py's _get_client()
is first called.
"""

from dotenv import load_dotenv
load_dotenv()

import asyncio
import json
import logging
import os
import tempfile
from pathlib import Path
from urllib.parse import urlsplit
from typing import Dict, List, Optional, Union, cast
from typing_extensions import TypedDict

from openpyxl import load_workbook
from fastapi import FastAPI, File, HTTPException, Security, UploadFile
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.types import ASGIApp, Message, Receive, Scope, Send

from skill.assembler import AssemblyError
from skill.constants import DEFAULT_MISTRAL_MODEL, MAX_REQUEST_BODY_BYTES, MAX_UPLOAD_FILE_SIZE
from skill.errors import PipelineStepError
from skill.parser_utils import (
    ComplexCaseRow,
    SchemaFieldRow,
    TrainingCaseRow,
    parse_complex_case,
    parse_training_cases,
)
from skill.pipeline import run_pipeline
from skill.schema import PreAuthCaseInput, PreAuthSkillOutput

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Security: API key for protected endpoints
# In production, use a proper auth system (OAuth2, JWT, etc.)
API_KEY = os.environ.get("API_KEY", "")
bearer_scheme = HTTPBearer(auto_error=False)

# Concurrency control for validate-all endpoint
# Limits concurrent validation runs to prevent resource exhaustion
_validate_all_semaphore: Optional[asyncio.Semaphore] = None


def get_validate_all_semaphore() -> asyncio.Semaphore:
    """Get or create the semaphore for validate-all concurrency control."""
    global _validate_all_semaphore
    if _validate_all_semaphore is None:
        # Allow max 1 concurrent validation run
        _validate_all_semaphore = asyncio.Semaphore(1)
    return _validate_all_semaphore


async def verify_api_key(
    credentials: Optional[HTTPAuthorizationCredentials] = Security(bearer_scheme),
) -> str:
    """
    Verify API key for protected endpoints.
    
    In production, replace with proper authentication (OAuth2, JWT, etc.).
    For development, if API_KEY is not set, the endpoint is unprotected.
    """
    if not API_KEY:
        # No API key configured - allow access (development mode)
        return "dev"
    
    if credentials is None:
        raise HTTPException(
            status_code=401,
            detail="API key required. Provide via Authorization: Bearer <key>",
        )
    
    if credentials.credentials != API_KEY:
        raise HTTPException(
            status_code=401,
            detail="Invalid API key",
        )
    
    return credentials.credentials


class RequestBodyTooLarge(Exception):
    """Internal control-flow signal for request body size violations."""


class RequestSizeLimitMiddleware:
    """Reject requests whose bodies exceed a configured maximum size."""

    def __init__(self, app: ASGIApp, max_bytes: int) -> None:
        self.app = app
        self.max_bytes = max_bytes

    async def __call__(self, scope: Scope, receive: Receive, send: Send) -> None:
        if scope["type"] != "http":
            await self.app(scope, receive, send)
            return

        content_length: Optional[int] = None
        for header_name, header_value in scope.get("headers", []):
            if header_name == b"content-length":
                try:
                    content_length = int(header_value.decode("latin-1"))
                except ValueError:
                    content_length = None
                break

        if content_length is not None and content_length > self.max_bytes:
            await JSONResponse(
                {
                    "detail": (
                        f"Request body too large. Maximum allowed size is {self.max_bytes} bytes."
                    )
                },
                status_code=413,
            )(scope, receive, send)
            return

        bytes_seen = 0

        async def limited_receive() -> Message:
            nonlocal bytes_seen
            message = await receive()
            if message["type"] == "http.request":
                bytes_seen += len(message.get("body", b""))
                if bytes_seen > self.max_bytes:
                    raise RequestBodyTooLarge
            return message

        try:
            await self.app(scope, limited_receive, send)
        except RequestBodyTooLarge:
            response = JSONResponse(
                {
                    "detail": (
                        f"Request body too large. Maximum allowed size is {self.max_bytes} bytes."
                    )
                },
                status_code=413,
            )
            await response(scope, receive, send)
            return


class ValidationResult(TypedDict):
    """Type for a single validation result in validate_all."""
    case_id: str
    requested_service: Optional[str]
    expected: str
    actual: str
    match: bool
    confidence: Optional[str]
    complexity_notes: Optional[str]
    note: Optional[str]


class ValidationSummary(TypedDict):
    """Type for the validate_all response."""
    total: int
    correct: int
    accuracy: float
    results: List[ValidationResult]


DATA_DIR = Path(__file__).parent / "data"


def _normalize_origin(origin: str) -> Optional[str]:
    """Return a normalized origin string or None for invalid input."""
    candidate = origin.strip().rstrip("/")
    if not candidate:
        return None

    parsed = urlsplit(candidate)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        logger.warning("Ignoring invalid CORS origin: %s", origin)
        return None

    return f"{parsed.scheme}://{parsed.netloc}"


def _get_allowed_origins() -> List[str]:
    """Collect and normalize allowed CORS origins from environment variables."""
    origin_sources = [
        os.environ.get("ALLOWED_ORIGINS", ""),
        os.environ.get("FRONTEND_URL", ""),
        os.environ.get("NEXT_PUBLIC_FRONTEND_URL", ""),
    ]

    origins: List[str] = []
    seen: set[str] = set()
    for source in origin_sources:
        for raw_origin in source.split(","):
            normalized = _normalize_origin(raw_origin)
            if normalized is None or normalized in seen:
                continue
            seen.add(normalized)
            origins.append(normalized)

    logger.info("Configured CORS allowed origins: %s", origins)
    return origins

app = FastAPI(
    title="Pre-Auth Copilot API",
    description="LLM-powered pre-authorization medical necessity reviewer",
    version="1.0.0",
)

app.add_middleware(RequestSizeLimitMiddleware, max_bytes=MAX_REQUEST_BODY_BYTES)

# CORS — declared before routes. Change allow_origins to deployed domain in production.
origins_list = _get_allowed_origins()

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins_list,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_json(filename: str) -> object:
    path = DATA_DIR / filename
    if not path.exists():
        raise HTTPException(
            status_code=503,
            detail=f"Data file not found: {filename}. Run scripts/parse_workbook.py first.",
        )
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _load_training_cases() -> List[TrainingCaseRow]:
    return cast(List[TrainingCaseRow], _load_json("training_cases.json"))


def _load_schema_fields() -> List[SchemaFieldRow]:
    return cast(List[SchemaFieldRow], _load_json("patient_data_schema.json"))


def _require_training_case_text(
    value: Optional[str],
    field_name: str,
    case_id: Optional[str] = None,
) -> str:
    if value is None or value == "":
        if case_id is None:
            raise HTTPException(
                status_code=500,
                detail=f"Training case is missing required field: {field_name}",
            )
        raise HTTPException(
            status_code=500,
            detail=f"Training case {case_id} is missing required field: {field_name}",
        )
    return value


def _complex_case_rows_to_input(rows: List[ComplexCaseRow]) -> PreAuthCaseInput:
    """Map the 19-row complex case format to a PreAuthCaseInput."""
    from skill.schema import SourcedField
    field_map: Dict[str, ComplexCaseRow] = {
        str(r.get("field", "")).lower().strip(): r
        for r in rows if r.get("field")
    }

    def _get_value(name: str) -> Optional[str]:
        row = field_map.get(name.lower())
        return row.get("value") if row else None

    def _get_row(name: str) -> Optional[ComplexCaseRow]:
        return field_map.get(name.lower())

    def _get_sourced(name: str) -> Optional[SourcedField]:
        row = field_map.get(name.lower())
        if not row or not row.get("value"):
            return None
        return SourcedField(value=row.get("value"), source=row.get("source"), source_date=None)

    age_raw = _get_value("age / sex") or _get_value("age")
    age = None
    sex = None
    if age_raw:
        parts = [p.strip() for p in str(age_raw).split("/")]
        if parts and parts[0]:
            try:
                age = int(parts[0])
            except (ValueError, IndexError):
                pass
        if len(parts) > 1 and parts[1]:
            sex = parts[1]

    secondary_raw = _get_value("secondary diagnoses") or _get_value("comorbidities")
    secondary_diagnoses: Optional[List[str]] = None
    if secondary_raw:
        parts = [p.strip() for p in str(secondary_raw).replace(";", ",").split(",") if p.strip()]
        secondary_diagnoses = parts if parts else None

    conservative_care = _get_value("conservative care")
    prior_treatment: Optional[List[str]] = [conservative_care] if conservative_care else None
    requested_service_row = _get_row("requested service") or {}
    requested_service_notes = str(requested_service_row.get("notes") or "")
    site_of_care = _get_value("site of care") or _get_value("site_of_care")
    if not site_of_care and "inpatient" in requested_service_notes.lower():
        site_of_care = "Inpatient"
    requested_los = _get_value("expected los")
    if not requested_los and "2 midnights" in requested_service_notes.lower():
        requested_los = "2 midnights"
    risk_factors = _get_value("medical risk factors")
    if not secondary_diagnoses and risk_factors:
        secondary_diagnoses = [
            part.strip()
            for part in str(risk_factors).replace(";", ",").split(",")
            if part.strip()
        ]

    case_id_from_rows = _get_value("case_id") or "PA-001"
    return PreAuthCaseInput(
        case_id=case_id_from_rows,
        requested_service=_get_value("requested service") or "Unknown",
        site_of_care=site_of_care,
        requested_los=requested_los,
        payer_policy_version=_get_value("payer policy used") or _get_value("payer policy"),
        age=age,
        sex=sex,
        primary_diagnosis=_get_value("primary diagnosis") or "Unknown",
        secondary_diagnoses=secondary_diagnoses,
        symptom_duration=_get_value("symptom duration"),
        pain_severity=_get_value("pain severity"),
        functional_impairment_adls=_get_sourced("adl impact"),
        objective_neurologic_deficits=_get_sourced("objective findings"),
        imaging_findings=_get_sourced("mri cervical spine") or _get_sourced("imaging"),
        prior_conservative_treatment=prior_treatment,
        response_to_prior_treatment=_get_value("response"),
        complications_red_flags=_get_value("recent events"),
        contradictory_flags=_get_value("contradiction"),
        missing_records=_get_value("missing documentation"),
        ordering_provider_specialty=_get_value("ordering provider specialty"),
    )


async def _run_pipeline_or_raise_http(
    case_input: PreAuthCaseInput,
) -> PreAuthSkillOutput:
    try:
        return await run_pipeline(case_input)
    except AssemblyError as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "assembly_failed",
                "step": "step3",
                "validation_error": exc.validation_error,
                "raw_eval_snippet": exc.raw_eval_snippet,
            },
        ) from exc
    except PipelineStepError as exc:
        detail: Dict[str, object] = {
            "error": "pipeline_failed",
            "step": exc.step,
            "message": exc.message,
        }
        if exc.raw_llm_snippet is not None:
            detail["raw_llm_snippet"] = exc.raw_llm_snippet
        raise HTTPException(status_code=422, detail=detail) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.get("/api/health")
async def health() -> Dict[str, object]:
    return {
        "status": "ok",
        "model": DEFAULT_MISTRAL_MODEL,
    }


@app.post("/api/analyze", response_model=PreAuthSkillOutput)
async def analyze(case_input: PreAuthCaseInput) -> PreAuthSkillOutput:
    """Main route. Accepts PreAuthCaseInput JSON, returns PreAuthSkillOutput JSON."""
    return await _run_pipeline_or_raise_http(case_input)


@app.post("/api/analyze/upload", response_model=PreAuthSkillOutput)
async def analyze_upload(file: UploadFile = File(...)) -> PreAuthSkillOutput:
    """
    Accepts .xlsx upload, parses it, builds PreAuthCaseInput, runs pipeline.
    Imports parse_complex_case from skill.parser_utils — top-level import,
    fails at startup if missing (not silently at request time).
    """
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    # Check file size before reading into memory
    file_size = 0
    chunk_size = 8192
    chunks: List[bytes] = []
    while True:
        chunk = await file.read(chunk_size)
        if not chunk:
            break
        file_size += len(chunk)
        if file_size > MAX_UPLOAD_FILE_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum allowed size is {MAX_UPLOAD_FILE_SIZE} bytes.",
            )
        chunks.append(chunk)
    content = b"".join(chunks)

    tmp_path: Optional[str] = None
    workbook = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
            tmp.write(content)

        workbook = load_workbook(tmp_path, read_only=True, data_only=True)
        if "Complex_Case_Input" in workbook.sheetnames:
            rows = parse_complex_case(workbook)
            case_input = _complex_case_rows_to_input(rows)
        elif "Training_Cases" in workbook.sheetnames:
            cases = parse_training_cases(workbook)
            if not cases:
                raise HTTPException(status_code=422, detail="No cases found in Training_Cases sheet.")
            tc = cases[0]
            case_input = PreAuthCaseInput(
                case_id=tc.get("case_id") or "Unknown",
                requested_service=tc.get("requested_service") or "Unknown",
                primary_diagnosis=tc.get("clinical_scenario") or "Unknown",
                raw_clinical_notes=tc.get("key_supporting_evidence") or None,
            )
        else:
            raise HTTPException(
                status_code=422,
                detail="Could not find a recognizable case sheet in the uploaded workbook.",
            )

        return await _run_pipeline_or_raise_http(case_input)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                logger.exception("Failed to close uploaded workbook.")
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except FileNotFoundError:
                pass
            except OSError:
                logger.exception("Failed to delete temporary workbook %s", tmp_path)
        try:
            await file.close()
        except Exception:
            logger.exception("Failed to close uploaded file.")


@app.get("/api/cases")
async def get_cases() -> List[TrainingCaseRow]:
    """Returns all 10 training cases."""
    return _load_training_cases()


@app.get("/api/cases/{case_id}")
async def get_case(case_id: str) -> Dict[str, object]:
    """
    Returns single training case.
    For PA-001, also returns has_full_packet: true so the frontend can
    offer 'Load Full Clinical Packet' option.
    """
    cases = _load_training_cases()
    for case in cases:
        if case.get("case_id") == case_id:
            result = dict(case)
            result["has_full_packet"] = (case_id == "PA-001")
            return result
    raise HTTPException(status_code=404, detail=f"Case {case_id} not found.")


@app.get("/api/schema")
async def get_schema() -> List[SchemaFieldRow]:
    """Returns 30-field patient data schema with why_it_matters tooltips."""
    return _load_schema_fields()


@app.get("/api/schema/json")
async def get_schema_json() -> Dict[str, object]:
    """Returns machine-readable JSON Schema (D7) generated from Pydantic."""
    return {
        "input": PreAuthCaseInput.model_json_schema(),
        "output": PreAuthSkillOutput.model_json_schema(),
    }


@app.get("/api/complex-case/input")
async def get_complex_case_input() -> PreAuthCaseInput:
    """Returns the PreAuthCaseInput for the full complex case (PA-001)."""
    rows = _load_json("complex_case.json")
    return _complex_case_rows_to_input(cast(List[ComplexCaseRow], rows))


@app.get("/api/validate-all")
async def validate_all(
    _: str = Security(verify_api_key),
) -> ValidationSummary:
    """
    Runs all 10 training cases through the pipeline.
    Returns accuracy table. Saves full output per case to outputs/case_results/.
    
    Requires API key authentication. Only one validation run can execute at a time
    to prevent resource exhaustion and excessive API costs.
    """
    semaphore = get_validate_all_semaphore()
    
    async with semaphore:
        cases: List[TrainingCaseRow] = _load_training_cases()
        outputs_dir = Path(__file__).parent.parent / "outputs" / "case_results"
        outputs_dir.mkdir(parents=True, exist_ok=True)

        results: List[ValidationResult] = []
        for tc in cases:
            case_id = _require_training_case_text(tc.get("case_id"), "case_id")
            requested_service = _require_training_case_text(
                tc.get("requested_service"),
                "requested_service",
                case_id,
            )
            expected_outcome = _require_training_case_text(
                tc.get("expected_outcome"),
                "expected_outcome",
                case_id,
            )
            clinical_scenario = tc.get("clinical_scenario") or "See clinical scenario"
            key_supporting_evidence = tc.get("key_supporting_evidence") or ""
            key_gaps_or_risks = tc.get("key_gaps_or_risks") or ""
            complexity_notes = tc.get("complexity_notes")

            case_input = PreAuthCaseInput(
                case_id=case_id,
                requested_service=requested_service,
                primary_diagnosis=clinical_scenario,
                raw_clinical_notes=(
                    f"Clinical scenario: {clinical_scenario}\n"
                    f"Supporting evidence: {key_supporting_evidence}\n"
                    f"Gaps or risks: {key_gaps_or_risks}"
                ),
            )
            try:
                output = await run_pipeline(case_input)
                match = output.recommendation == expected_outcome
                out_path = outputs_dir / f"{case_id}.json"
                with open(out_path, "w") as f:
                    json.dump(output.model_dump(), f, indent=2)
                result: ValidationResult = {
                    "case_id": case_id,
                    "requested_service": requested_service,
                    "expected": expected_outcome,
                    "actual": output.recommendation,
                    "match": match,
                    "confidence": output.confidence,
                    "complexity_notes": complexity_notes,
                    "note": None if match else f"Expected {expected_outcome}, got {output.recommendation}",
                }
                results.append(result)
            except Exception as e:
                error_result: ValidationResult = {
                    "case_id": case_id,
                    "requested_service": None,
                    "expected": expected_outcome,
                    "actual": "ERROR",
                    "match": False,
                    "confidence": None,
                    "complexity_notes": complexity_notes,
                    "note": str(e),
                }
                results.append(error_result)

        correct = sum(1 for r in results if r["match"])
        return {
            "total": len(results),
            "correct": correct,
            "accuracy": round(correct / len(results), 2) if results else 0,
            "results": results,
        }


class ProgressEvent(TypedDict):
    """Type for progress update events."""
    type: str
    current: int
    total: int
    case_id: Optional[str]
    result: Optional[Union[ValidationResult, ValidationSummary]]


@app.get("/api/validate-all/stream")
async def validate_all_stream(
    _: str = Security(verify_api_key),
) -> StreamingResponse:
    """
    Server-Sent Events endpoint for real-time validation progress.
    Returns progress updates as each case is processed.
    """
    
    async def event_generator():
        semaphore = get_validate_all_semaphore()
        
        async with semaphore:
            cases: List[TrainingCaseRow] = _load_training_cases()
            outputs_dir = Path(__file__).parent.parent / "outputs" / "case_results"
            outputs_dir.mkdir(parents=True, exist_ok=True)

            results: List[ValidationResult] = []
            for i, tc in enumerate(cases):
                case_id = _require_training_case_text(tc.get("case_id"), "case_id")
                requested_service = _require_training_case_text(
                    tc.get("requested_service"),
                    "requested_service",
                    case_id,
                )
                expected_outcome = _require_training_case_text(
                    tc.get("expected_outcome"),
                    "expected_outcome",
                    case_id,
                )
                clinical_scenario = tc.get("clinical_scenario") or "See clinical scenario"
                key_supporting_evidence = tc.get("key_supporting_evidence") or ""
                key_gaps_or_risks = tc.get("key_gaps_or_risks") or ""
                complexity_notes = tc.get("complexity_notes")

                case_input = PreAuthCaseInput(
                    case_id=case_id,
                    requested_service=requested_service,
                    primary_diagnosis=clinical_scenario,
                    raw_clinical_notes=(
                        f"Clinical scenario: {clinical_scenario}\n"
                        f"Supporting evidence: {key_supporting_evidence}\n"
                        f"Gaps or risks: {key_gaps_or_risks}"
                    ),
                )
                
                # Send progress event
                progress: ProgressEvent = {
                    "type": "progress",
                    "current": i + 1,
                    "total": len(cases),
                    "case_id": case_id,
                    "result": None,
                }
                yield f"data: {json.dumps(progress)}\n\n"
                
                try:
                    output = await run_pipeline(case_input)
                    match = output.recommendation == expected_outcome
                    out_path = outputs_dir / f"{case_id}.json"
                    with open(out_path, "w") as f:
                        json.dump(output.model_dump(), f, indent=2)
                    result: ValidationResult = {
                        "case_id": case_id,
                        "requested_service": requested_service,
                        "expected": expected_outcome,
                        "actual": output.recommendation,
                        "match": match,
                        "confidence": output.confidence,
                        "complexity_notes": complexity_notes,
                        "note": None if match else f"Expected {expected_outcome}, got {output.recommendation}",
                    }
                    results.append(result)
                except Exception as e:
                    error_result: ValidationResult = {
                        "case_id": case_id,
                        "requested_service": None,
                        "expected": expected_outcome,
                        "actual": "ERROR",
                        "match": False,
                        "confidence": None,
                        "complexity_notes": complexity_notes,
                        "note": str(e),
                    }
                    results.append(error_result)

            # Send final result
            correct = sum(1 for r in results if r["match"])
            final: ProgressEvent = {
                "type": "complete",
                "current": len(results),
                "total": len(cases),
                "case_id": None,
                "result": {
                    "total": len(results),
                    "correct": correct,
                    "accuracy": round(correct / len(results), 2) if results else 0,
                    "results": results,
                },
            }
            yield f"data: {json.dumps(final)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        },
    )
